#!/usr/bin/env python3
"""
Timekeeper — free, open-source, ad-free macOS time tracker.
Run:  python timekeeper_app.py
Deps: pip install PySide6 openpyxl pyobjc-framework-Cocoa
"""

# ── Imports, Configuration, Sound Helper & First-launch Dir Picker ────────────
import sys, os, uuid, sqlite3, subprocess, ctypes
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from typing import List, Optional
from pathlib import Path

from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QPushButton, QInputDialog, QMessageBox, QFileDialog, QMenu,
    QDialog, QListWidget, QListWidgetItem, QDialogButtonBox
)
from PySide6.QtCore import Qt, QTimer, Signal, QObject
from PySide6.QtGui import QFont, QColor, QPainter, QPen

import openpyxl
from openpyxl.styles import Font as XLFont, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Configuration ─────────────────────────────────────────────────────────────
APP_NAME           = "Timekeeper"
APP_VERSION        = "v0.00006"
CONFIG_PATH        = Path.home() / ".timekeeper_config"
DEFAULT_TASK_MINS  = 30
DEFAULT_BREAK_MINS = 5
DEFAULT_DAILY_GOAL = 10

# ── First-launch data directory picker ────────────────────────────────────────
def get_or_create_data_dir() -> Path:
    """Return the data directory. On first launch, ask the user to choose one."""
    if CONFIG_PATH.exists():
        saved = Path(CONFIG_PATH.read_text().strip())
        saved.mkdir(parents=True, exist_ok=True)
        return saved
    QMessageBox.information(
        None, "Welcome to Timekeeper",
        "Please choose a folder where Timekeeper will store your data\n"
        "(database and Excel timesheet).\n\n"
        "You can move this folder later by editing ~/.timekeeper_config"
    )
    folder = QFileDialog.getExistingDirectory(
        None, "Choose data folder", str(Path.home() / "Documents")
    )
    data_dir = Path(folder) if folder else Path.home() / "Documents" / "Timekeeper"
    data_dir.mkdir(parents=True, exist_ok=True)
    CONFIG_PATH.write_text(str(data_dir))
    return data_dir

# ── Sound helper ──────────────────────────────────────────────────────────────
def play_sound(name: str, rate: float = 1.0, volume: float = 1.0) -> None:
    """Play a named macOS system sound via afplay. Fails silently."""
    sounds = {
        "Glass": "/System/Library/Sounds/Glass.aiff",
        "Funk":  "/System/Library/Sounds/Funk.aiff",
        "Tink":  "/System/Library/Sounds/Tink.aiff",
    }
    path = sounds.get(name)
    if path:
        # [DEBUG D1] – remove before publish
        print(f"[DEBUG] play_sound: {name} at {datetime.now().strftime('%H:%M:%S')}")
        cmd = ["afplay"]
        if rate != 1.0:
            cmd += ["-r", str(rate)]
        if volume != 1.0:
            cmd += ["-v", str(volume)]
        cmd.append(path)
        subprocess.Popen(cmd)


# ── Data Models ───────────────────────────────────────────────────────────────
@dataclass
class Task:
    name: str
    id: str                    = field(default_factory=lambda: str(uuid.uuid4()))
    linked_path: Optional[str] = None
    created_at: datetime       = field(default_factory=datetime.now)

    def __str__(self) -> str:
        return self.name


@dataclass
class TimeEntry:
    task_id:        str
    task_name:      str
    start_time:     datetime
    end_time:       Optional[datetime] = None
    manually_added: bool               = False
    notes:          str                = ""
    interval_id:    Optional[str]      = None

    @property
    def duration_minutes(self) -> float:
        if self.end_time:
            return (self.end_time - self.start_time).total_seconds() / 60
        return 0.0


# ── Storage ───────────────────────────────────────────────────────────────────
class IStorage(ABC):
    @abstractmethod
    def save_task(self, task: Task) -> None: ...
    @abstractmethod
    def get_tasks(self) -> List[Task]: ...
    @abstractmethod
    def update_task(self, task: Task) -> None: ...
    @abstractmethod
    def delete_task(self, task_id: str) -> None: ...
    @abstractmethod
    def save_time_entry(self, entry: TimeEntry) -> None: ...
    @abstractmethod
    def get_time_entries(self) -> List[TimeEntry]: ...
    @abstractmethod
    def get_today_count(self) -> int: ...
    @abstractmethod
    def get_setting(self, key: str) -> Optional[str]: ...
    @abstractmethod
    def set_setting(self, key: str, value: str) -> None: ...
    @abstractmethod
    def get_next_interval_id(self) -> str: ...


class SQLiteStorage(IStorage):
    def __init__(self, db_path: str):
        self.db_path = db_path
        os.makedirs(os.path.dirname(db_path), exist_ok=True)
        self._init_db()

    def _init_db(self) -> None:
        with sqlite3.connect(self.db_path) as conn:
            conn.execute('''
                CREATE TABLE IF NOT EXISTS tasks (
                    id TEXT PRIMARY KEY, name TEXT NOT NULL,
                    linked_path TEXT, created_at TEXT NOT NULL)''')
            conn.execute('''
                CREATE TABLE IF NOT EXISTS time_entries (
                    id TEXT PRIMARY KEY, task_id TEXT NOT NULL,
                    task_name TEXT NOT NULL, start_time TEXT NOT NULL,
                    end_time TEXT, manually_added INTEGER DEFAULT 0,
                    notes TEXT DEFAULT "", interval_id TEXT)''')
            conn.execute('''
                CREATE TABLE IF NOT EXISTS settings (
                    key TEXT PRIMARY KEY, value TEXT NOT NULL)''')

    def save_task(self, task: Task) -> None:
        with sqlite3.connect(self.db_path) as conn:
            conn.execute('INSERT OR REPLACE INTO tasks VALUES (?,?,?,?)',
                (task.id, task.name, task.linked_path, task.created_at.isoformat()))

    def get_tasks(self) -> List[Task]:
        with sqlite3.connect(self.db_path) as conn:
            rows = conn.execute('SELECT * FROM tasks ORDER BY created_at').fetchall()
        return [Task(name=r[1], id=r[0], linked_path=r[2],
                     created_at=datetime.fromisoformat(r[3])) for r in rows]

    def update_task(self, task: Task) -> None:
        with sqlite3.connect(self.db_path) as conn:
            conn.execute('UPDATE tasks SET name=?, linked_path=? WHERE id=?',
                (task.name, task.linked_path, task.id))

    def delete_task(self, task_id: str) -> None:
        with sqlite3.connect(self.db_path) as conn:
            conn.execute('DELETE FROM tasks WHERE id=?', (task_id,))

    def save_time_entry(self, entry: TimeEntry) -> None:
        with sqlite3.connect(self.db_path) as conn:
            conn.execute('INSERT INTO time_entries VALUES (?,?,?,?,?,?,?,?)',
                (str(uuid.uuid4()), entry.task_id, entry.task_name,
                 entry.start_time.isoformat(),
                 entry.end_time.isoformat() if entry.end_time else None,
                 int(entry.manually_added), entry.notes, entry.interval_id))

    def get_time_entries(self) -> List[TimeEntry]:
        with sqlite3.connect(self.db_path) as conn:
            rows = conn.execute('SELECT * FROM time_entries ORDER BY start_time').fetchall()
        return [TimeEntry(task_id=r[1], task_name=r[2],
                          start_time=datetime.fromisoformat(r[3]),
                          end_time=datetime.fromisoformat(r[4]) if r[4] else None,
                          manually_added=bool(r[5]), notes=r[6] or "",
                          interval_id=r[7] if len(r) > 7 else None) for r in rows]

    def get_today_count(self) -> int:
        today = datetime.now().date().isoformat()
        with sqlite3.connect(self.db_path) as conn:
            row = conn.execute(
                "SELECT COUNT(DISTINCT interval_id) FROM time_entries "
                "WHERE manually_added=0 AND interval_id IS NOT NULL "
                "AND interval_id NOT LIKE '%-999' AND substr(start_time,1,10)=?",
                (today,)
            ).fetchone()
        return row[0] if row else 0

    def get_setting(self, key: str) -> Optional[str]:
        with sqlite3.connect(self.db_path) as conn:
            row = conn.execute('SELECT value FROM settings WHERE key=?', (key,)).fetchone()
        return row[0] if row else None

    def set_setting(self, key: str, value: str) -> None:
        with sqlite3.connect(self.db_path) as conn:
            conn.execute('INSERT OR REPLACE INTO settings VALUES (?,?)', (key, value))

    def get_next_interval_id(self) -> str:
        today = datetime.now().strftime("%Y%m%d")
        with sqlite3.connect(self.db_path) as conn:
            rows = conn.execute(
                "SELECT interval_id FROM time_entries WHERE interval_id LIKE ?",
                (today + "-%",)
            ).fetchall()
        max_seq = 0
        for (iid,) in rows:
            try:
                seq = int(iid.split("-")[1])
                if seq < 999:
                    max_seq = max(max_seq, seq)
            except (IndexError, ValueError):
                pass
        return f"{today}-{max_seq + 1:03d}"


# ── Timer Engine ──────────────────────────────────────────────────────────────
class TimerEngine(QObject):
    tick      = Signal(int, int)
    completed = Signal()

    def __init__(self):
        super().__init__()
        self._qtimer = QTimer()
        self._qtimer.setInterval(1000)
        self._qtimer.timeout.connect(self._tick)
        self.seconds_remaining: int  = 0
        self.total_seconds:     int  = 0
        self.is_running:        bool = False
        self.is_break:          bool = False

    def start(self, seconds: int, is_break: bool = False) -> None:
        self.stop()
        self.total_seconds     = seconds
        self.seconds_remaining = seconds
        self.is_break          = is_break
        self.is_running        = True
        self._qtimer.start()
        self.tick.emit(self.seconds_remaining, self.total_seconds)

    def pause(self) -> None:
        self._qtimer.stop()
        self.is_running = False

    def resume(self) -> None:
        if self.seconds_remaining > 0:
            self._qtimer.start()
            self.is_running = True

    def stop(self) -> None:
        self._qtimer.stop()
        self.is_running        = False
        self.seconds_remaining = 0
        self.total_seconds     = 0

    @staticmethod
    def format_time(seconds: int) -> str:
        m, s = divmod(abs(seconds), 60)
        h, m = divmod(m, 60)
        return f"{h:02d}:{m:02d}:{s:02d}" if h else f"{m:02d}:{s:02d}"

    def _tick(self) -> None:
        self.seconds_remaining -= 1
        self.tick.emit(self.seconds_remaining, self.total_seconds)
        if self.seconds_remaining <= 0:
            self._qtimer.stop()
            self.is_running = False
            self.completed.emit()


# ── Excel Exporter ────────────────────────────────────────────────────────────
class ExcelExporter:
    _HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
    _HEADER_FONT = XLFont(bold=True, color="FFFFFF", size=11)
    _ALT_FILL    = PatternFill("solid", fgColor="EBF5FB")
    _LINK_FONT   = XLFont(color="0563C1", underline="single")

    def __init__(self, filepath: str):
        self.filepath = filepath
        os.makedirs(os.path.dirname(filepath), exist_ok=True)

    def export(self, entries: List[TimeEntry], tasks: List[Task]) -> None:
        wb = openpyxl.Workbook()
        self._write_time_log(wb, entries)
        self._write_task_links(wb, tasks)
        wb.save(self.filepath)

    def _write_time_log(self, wb, entries: List[TimeEntry]) -> None:
        ws = wb.active
        ws.title = "Time Log"
        headers = ["Date", "Task Name", "Interval ID", "Duration (min)", "Start", "End", "Manual Entry"]
        self._write_headers(ws, headers, [12, 30, 15, 15, 10, 10, 13])
        for i, e in enumerate(entries, 2):
            row = [e.start_time.strftime("%Y-%m-%d"), e.task_name,
                   e.interval_id or "",
                   round(e.duration_minutes, 1), e.start_time.strftime("%H:%M:%S"),
                   e.end_time.strftime("%H:%M:%S") if e.end_time else "—",
                   "Yes" if e.manually_added else "No"]
            for col, val in enumerate(row, 1):
                cell = ws.cell(i, col, val)
                if i % 2 == 0:
                    cell.fill = self._ALT_FILL

    def _write_task_links(self, wb, tasks: List[Task]) -> None:
        ws = wb.create_sheet("Task Links")
        self._write_headers(ws, ["Task Name", "Linked File / Folder", "Date Created"],
                            [30, 60, 15])
        for i, t in enumerate([t for t in tasks if t.linked_path], 2):
            ws.cell(i, 1, t.name)
            cell = ws.cell(i, 2, t.linked_path)
            cell.hyperlink = f"file://{t.linked_path}"
            cell.font = self._LINK_FONT
            ws.cell(i, 3, t.created_at.strftime("%Y-%m-%d"))

    def _write_headers(self, ws, headers: List[str], widths: List[int]) -> None:
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(1, col, h)
            cell.font = self._HEADER_FONT
            cell.fill = self._HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col)].width = w


# ── Circular Progress Widget ──────────────────────────────────────────────────
class CircularProgress(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMinimumSize(150, 150)
        self._fraction = 1.0
        self._is_break = False

    def set_value(self, fraction: float, is_break: bool = False) -> None:
        self._fraction = max(0.0, min(1.0, fraction))
        self._is_break = is_break
        self.update()

    def paintEvent(self, _event) -> None:
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        r    = self.rect()
        size = min(r.width(), r.height()) - 20
        x    = (r.width()  - size) // 2
        y    = (r.height() - size) // 2
        arc  = (x + 8, y + 8, size - 16, size - 16)
        painter.setPen(QPen(QColor("#2C3E50"), 10))
        painter.setBrush(Qt.NoBrush)
        painter.drawEllipse(*arc)
        if self._is_break:
            colour = QColor("#F39C12")
        elif self._fraction > 0.5:
            colour = QColor("#27AE60")
        elif self._fraction > 0.25:
            colour = QColor("#F39C12")
        else:
            colour = QColor("#E74C3C")
        painter.setPen(QPen(colour, 10, Qt.SolidLine, Qt.RoundCap))
        painter.drawArc(*arc, 90 * 16, int(self._fraction * 360 * 16))
        painter.end()


# ── Styles & Constants ────────────────────────────────────────────────────────
STYLE = """
QWidget          { background:#1E2329; color:#ECEFF1;
                   font-family:'Helvetica Neue',Arial; font-size:13px; }
QPushButton      { background:#2C3E50; color:#ECEFF1; border:none;
                   padding:7px 14px; border-radius:6px; }
QPushButton:hover{ background:#34495E; }
QPushButton:disabled { background:#1A2029; color:#4A5568; }
QPushButton#go   { background:#27AE60; }
QPushButton#go:hover  { background:#2ECC71; }
QPushButton#stop { background:#C0392B; }
QPushButton#stop:hover{ background:#E74C3C; }
QDialog          { background:#1E2329; color:#ECEFF1; }
QListView        { background:#2C3E50; color:#ECEFF1; border:none; }
QListView::item:selected { background:#27AE60; color:#ECEFF1; }
QListView::item:hover    { background:#34495E; color:#ECEFF1; }
QLineEdit        { background:#2C3E50; color:#ECEFF1; border:none;
                   padding:4px; border-radius:4px; }
QMenu            { background:#2C3E50; color:#ECEFF1;
                   border:1px solid #34495E; padding:4px; }
QMenu::item      { padding:6px 22px; }
QMenu::item:selected { background:#27AE60; color:#ECEFF1; }
QMenu::separator { height:1px; background:#34495E; margin:4px 2px; }
"""

CLR_TASK  = "#ECEFF1"
CLR_BREAK = "#F39C12"
CLR_READY = "#27AE60"
CLR_DIM   = "#7F8C8D"

CIRCLE_SIZE = 200


# ── Task Picker Dialog ────────────────────────────────────────────────────────
class TaskPickerDialog(QDialog):
    def __init__(self, tasks, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Switch Task")
        self.setMinimumWidth(300)
        self.setStyleSheet(STYLE)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(10)
        lbl = QLabel("Select task:")
        lbl.setStyleSheet(f"color:{CLR_DIM}; font-size:12px;")
        layout.addWidget(lbl)
        self._list = QListWidget()
        self._list.setStyleSheet(
            "QListWidget { background:#2C3E50; color:#ECEFF1; border:none; }"
            "QListWidget::item:selected { background:#27AE60; color:#ECEFF1; }"
            "QListWidget::item:hover    { background:#34495E; color:#ECEFF1; }"
        )
        for t in tasks:
            self._list.addItem(QListWidgetItem(t.name))
        self._list.setCurrentRow(0)
        self._list.itemDoubleClicked.connect(self.accept)
        layout.addWidget(self._list)
        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.setStyleSheet(
            "QPushButton { background:#2C3E50; color:#ECEFF1; border:none; "
            "padding:6px 14px; border-radius:5px; } "
            "QPushButton:hover { background:#34495E; }"
        )
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def selected_name(self):
        item = self._list.currentItem()
        return item.text() if item else None


# ── Main Panel UI ─────────────────────────────────────────────────────────────
class MainPanel(QWidget):
    def __init__(self, engine: TimerEngine, storage: IStorage,
                 exporter: ExcelExporter, excel_path: str):
        super().__init__()
        self._engine     = engine
        self._storage    = storage
        self._exporter   = exporter
        self._excel_path = excel_path

        self._current_task:       Optional[Task]     = None
        self._tasks:              List[Task]          = []
        self._task_period_start:  Optional[datetime] = None
        self._mode:               str                = "idle"
        self._interval_id:        Optional[str]      = None
        self._task_mins:          int                = DEFAULT_TASK_MINS
        self._break_mins:         int                = DEFAULT_BREAK_MINS
        self._daily_goal:         int                = DEFAULT_DAILY_GOAL

        self._ready = False
        self._build_ui()
        self._connect_signals()
        self._load_tasks()
        self._update_today_count()
        self._ready = True

    def _build_ui(self) -> None:
        self.setMinimumWidth(280)
        self.setWindowFlags(Qt.Window)
        self.setStyleSheet(STYLE)

        root = QVBoxLayout(self)
        root.setContentsMargins(12, 16, 12, 16)
        root.setSpacing(14)

        # Top row: hamburger menu (left) + task name (centred) + balancing spacer (right)
        task_row = QHBoxLayout()
        self._menu_btn = QPushButton("☰")
        self._menu_btn.setFixedWidth(36)
        self._menu_btn.setToolTip("Menu")
        self._menu_btn.clicked.connect(self._show_menu)
        task_row.addWidget(self._menu_btn)
        self._task_lbl = QLabel("No task — open menu to create one")
        self._task_lbl.setFont(QFont("Helvetica Neue", 16, QFont.Bold))
        self._task_lbl.setAlignment(Qt.AlignCenter)
        task_row.addWidget(self._task_lbl, 1)
        spacer = QWidget()
        spacer.setFixedWidth(36)
        task_row.addWidget(spacer)
        root.addLayout(task_row)

        # Today counter
        self._today_lbl = QLabel(f"Today: 0 / {DEFAULT_DAILY_GOAL} intervals")
        self._today_lbl.setAlignment(Qt.AlignCenter)
        self._today_lbl.setStyleSheet(f"color:{CLR_DIM}; font-size:12px; font-weight:bold;")
        root.addWidget(self._today_lbl)

        # Circular progress with clock overlaid inside
        self._progress = CircularProgress()
        self._progress.setFixedSize(CIRCLE_SIZE, CIRCLE_SIZE)

        self._clock_lbl = QLabel("00:00", self._progress)
        self._clock_lbl.setFont(QFont("Courier", 38, QFont.Bold))
        self._clock_lbl.setAlignment(Qt.AlignCenter)
        self._clock_lbl.setGeometry(0, 0, CIRCLE_SIZE, CIRCLE_SIZE)
        self._clock_lbl.setStyleSheet(
            f"color:{CLR_TASK}; letter-spacing:3px; background:transparent;")

        circle_row = QHBoxLayout()
        circle_row.addStretch()
        circle_row.addWidget(self._progress)
        circle_row.addStretch()
        root.addLayout(circle_row)

        # Interval length label below circle
        self._total_lbl = QLabel(f"Task length: {DEFAULT_TASK_MINS} minutes")
        self._total_lbl.setAlignment(Qt.AlignCenter)
        self._total_lbl.setStyleSheet(f"color:{CLR_DIM}; font-size:11px;")
        root.addWidget(self._total_lbl)

        # Controls — width matches circle diameter, centered
        ctrl_widget = QWidget()
        ctrl_widget.setFixedWidth(CIRCLE_SIZE)
        ctrl_layout = QHBoxLayout(ctrl_widget)
        ctrl_layout.setContentsMargins(0, 0, 0, 0)
        ctrl_layout.setSpacing(10)
        self._start_btn = QPushButton("▶  Start Task")
        self._start_btn.setObjectName("go")
        self._start_btn.clicked.connect(self._on_start_clicked)
        ctrl_layout.addWidget(self._start_btn)
        stop_btn = QPushButton("■  Stop")
        stop_btn.setObjectName("stop")
        stop_btn.clicked.connect(self._stop_timer)
        ctrl_layout.addWidget(stop_btn)

        ctrl_row = QHBoxLayout()
        ctrl_row.addStretch()
        ctrl_row.addWidget(ctrl_widget)
        ctrl_row.addStretch()
        root.addLayout(ctrl_row)

    def _show_about(self) -> None:
        QMessageBox.information(self, "About Timekeeper",
            f"Timekeeper {APP_VERSION}\n\n"
            "An open-source time tracking app for macOS (only, for now).\n\n"
            "Track tasks, record intervals, and exports to Excel."
        )

    def _show_shortcuts(self) -> None:
        QMessageBox.information(self, "Keyboard Shortcuts",
            "Ctrl + B    Start / Pause interval\n"
            "Ctrl + S    Stop interval\n"
            "Ctrl + M    Open menu - then navigate as usual"
        )

    def _show_menu(self) -> None:
        menu = QMenu(self)
        menu.addAction("＋  New Task",                self._new_task)
        menu.addAction("✏  Rename Task",             self._rename_task)
        menu.addAction("⇄  Switch Task",             self._switch_task)
        menu.addSeparator()
        menu.addAction("⏱  Task Interval…",    self._set_task_mins)
        menu.addAction("☕  Break Interval…",   self._set_break_mins)
        menu.addAction("◎  Daily Goal…",       self._set_daily_goal)
        menu.addSeparator()
        menu.addAction("＋  Add Time Manually",       self._add_time_manually)
        menu.addAction("\U0001f4ce  Add Link to Task",    self._link_file_or_folder)
        menu.addSeparator()
        menu.addAction("\U0001f5c4  Open Database",       self._open_db)
        menu.addAction("\U0001f4ca  Export & Open Excel", self._export_and_open_excel)
        menu.addSeparator()
        menu.addAction("ℹ  About",                       self._show_about)
        menu.addAction("⌨  Keyboard Shortcuts",          self._show_shortcuts)
        menu.addAction("Quit Timekeeper",                 self._quit)
        pos = self._menu_btn.mapToGlobal(self._menu_btn.rect().bottomLeft())
        menu.exec(pos)

    def _set_task_mins(self) -> None:
        val, ok = QInputDialog.getInt(self, "Task Interval",
                                      "Length (minutes):", self._task_mins, 1, 240)
        if ok:
            self._task_mins = val

    def _set_break_mins(self) -> None:
        val, ok = QInputDialog.getInt(self, "Break Interval",
                                      "Length (minutes):", self._break_mins, 1, 60)
        if ok:
            self._break_mins = val

    def _set_daily_goal(self) -> None:
        val, ok = QInputDialog.getInt(self, "Daily Goal",
                                      "Intervals per day:", self._daily_goal, 1, 50)
        if ok:
            self._daily_goal = val
            self._update_today_count()

    def _connect_signals(self) -> None:
        self._engine.tick.connect(self._on_tick)
        self._engine.completed.connect(self._on_complete)


    def _load_tasks(self) -> None:
        self._tasks = self._storage.get_tasks()
        if self._tasks:
            last_id = self._storage.get_setting('last_task_id')
            self._current_task = (
                next((t for t in self._tasks if t.id == last_id), None)
                or self._tasks[0]
            )
            self._task_lbl.setText(self._current_task.name)

    def _update_today_count(self) -> None:
        count = self._storage.get_today_count()
        self._today_lbl.setText(f"Today: {count} / {self._daily_goal} intervals")
        if count >= self._daily_goal:
            self._today_lbl.setStyleSheet(
                f"color:{CLR_READY}; font-size:12px; font-weight:bold;")
        else:
            self._today_lbl.setStyleSheet(
                f"color:{CLR_DIM}; font-size:12px; font-weight:bold;")

    def _on_tick(self, remaining: int, total: int) -> None:
        self._clock_lbl.setText(TimerEngine.format_time(remaining))
        fraction = remaining / total if total > 0 else 0.0
        self._progress.set_value(fraction, self._engine.is_break)

    def _on_complete(self) -> None:
        # [DEBUG D2] – remove before publish
        print(f"[DEBUG] _on_complete at {datetime.now().strftime('%H:%M:%S')}, "
              f"is_break={self._engine.is_break}, remaining={self._engine.seconds_remaining}")
        if self._engine.is_break:
            play_sound("Funk", volume=300)
            self._preload_task()
        else:
            play_sound("Glass", rate=0.2, volume=300)
            self._save_current_period(update_counter=True)
            self._preload_break()

    def _on_start_clicked(self) -> None:
        if self._mode == "break_ready":
            self._begin_break()
        elif self._mode == "task_ready":
            self._begin_task()
        elif self._mode == "task":
            self._engine.pause()
            self._mode = "task_paused"
            self._start_btn.setText("▶  Resume")
        elif self._mode == "task_paused":
            self._engine.resume()
            self._mode = "task"
            self._start_btn.setText("⏸  Pause")
        elif self._mode == "break":
            self._engine.pause()
            self._mode = "break_paused"
            self._start_btn.setText("▶  Resume")
        elif self._mode == "break_paused":
            self._engine.resume()
            self._mode = "break"
            self._start_btn.setText("⏸  Pause")
        else:
            self._begin_task()

    def _begin_task(self) -> None:
        if not self._current_task:
            QMessageBox.warning(self, "No Task", "Create or select a task first.")
            return
        self._update_today_count()
        self._storage.set_setting('last_task_id', self._current_task.id)
        if self._interval_id is None:
            self._interval_id = self._storage.get_next_interval_id()
        self._task_period_start = datetime.now()
        self._engine.start(self._task_mins * 60, is_break=False)
        self._mode = "task"
        self._start_btn.setText("⏸  Pause")
        self._clock_lbl.setStyleSheet(
            f"color:{CLR_TASK}; letter-spacing:3px; background:transparent;")
        self._total_lbl.setText(f"Task length: {self._task_mins} minutes")

    def _begin_break(self) -> None:
        self._engine.start(self._break_mins * 60, is_break=True)
        self._mode = "break"
        self._start_btn.setText("⏸  Pause")
        self._clock_lbl.setStyleSheet(
            f"color:{CLR_BREAK}; letter-spacing:3px; background:transparent;")
        self._task_lbl.setText("Break")
        self._total_lbl.setText(f"Break: {self._break_mins} minutes")

    def _stop_timer(self) -> None:
        if self._mode in ("task", "task_paused"):
            self._save_current_period(update_counter=True)
            self._preload_break()
        elif self._mode in ("break", "break_paused"):
            self._engine.stop()
            self._preload_task()

    def _preload_break(self) -> None:
        self._engine.stop()
        self._mode = "break_ready"
        self._start_btn.setText("▶  Start Break")
        secs = self._break_mins * 60
        self._clock_lbl.setText(TimerEngine.format_time(secs))
        self._clock_lbl.setStyleSheet(
            f"color:{CLR_BREAK}; letter-spacing:3px; background:transparent;")
        self._task_lbl.setText("Break")
        self._total_lbl.setText(f"Break: {self._break_mins} minutes")
        self._progress.set_value(1.0, is_break=True)

    def _preload_task(self) -> None:
        self._engine.stop()
        self._mode = "task_ready"
        self._start_btn.setText("▶  Start Task")
        secs = self._task_mins * 60
        self._clock_lbl.setText(TimerEngine.format_time(secs))
        self._clock_lbl.setStyleSheet(
            f"color:{CLR_TASK}; letter-spacing:3px; background:transparent;")
        task_name = self._current_task.name if self._current_task else "No task"
        self._task_lbl.setText(task_name)
        self._total_lbl.setText(f"Task length: {self._task_mins} minutes")
        self._progress.set_value(1.0, is_break=False)

    def _save_current_period(self, update_counter: bool = False) -> None:
        if self._task_period_start and self._current_task:
            self._storage.save_time_entry(TimeEntry(
                task_id=self._current_task.id,
                task_name=self._current_task.name,
                start_time=self._task_period_start,
                end_time=datetime.now(),
                interval_id=self._interval_id
            ))
            self._task_period_start = None
            if update_counter:
                self._interval_id = None
            self._refresh_excel()
            if update_counter:
                self._update_today_count()

    def _new_task(self) -> None:
        name, ok = QInputDialog.getText(self, "New Task", "Task name:")
        if ok and name.strip():
            if self._mode in ("task", "task_paused") and self._current_task:
                self._save_current_period()
                self._task_period_start = datetime.now()
            task = Task(name=name.strip())
            self._storage.save_task(task)
            self._tasks.append(task)
            self._current_task = task
            self._storage.set_setting('last_task_id', task.id)
            self._task_lbl.setText(task.name)

    def _rename_task(self) -> None:
        if not self._current_task:
            return
        name, ok = QInputDialog.getText(self, "Rename Task", "New name:",
                                        text=self._current_task.name)
        if ok and name.strip():
            self._current_task.name = name.strip()
            self._storage.update_task(self._current_task)
            self._task_lbl.setText(self._current_task.name)

    def _switch_task(self) -> None:
        if not self._tasks:
            QMessageBox.information(self, "No Tasks", "Create a task first.")
            return
        dlg = TaskPickerDialog(self._tasks, self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        name = dlg.selected_name()
        if not name:
            return
        new_task = next(t for t in self._tasks if t.name == name)
        if self._current_task and new_task.id == self._current_task.id:
            return
        if self._mode in ("task", "task_paused") and self._current_task:
            self._save_current_period()
            self._task_period_start = datetime.now()
        self._current_task = new_task
        self._storage.set_setting('last_task_id', new_task.id)
        self._task_lbl.setText(new_task.name)

    def _add_time_manually(self) -> None:
        if not self._current_task:
            QMessageBox.warning(self, "No Task", "Select a task first.")
            return
        mins, ok = QInputDialog.getInt(self, "Add Time", "Minutes to add:", 30, 1, 600)
        if ok:
            now = datetime.now()
            self._storage.save_time_entry(TimeEntry(
                task_id=self._current_task.id, task_name=self._current_task.name,
                start_time=now - timedelta(minutes=mins), end_time=now,
                manually_added=True, notes="Manually added",
                interval_id=datetime.now().strftime("%Y%m%d") + "-999"
            ))
            self._refresh_excel()
            QMessageBox.information(self, "Done",
                f"{mins} min added to '{self._current_task.name}'.")

    def _link_file_or_folder(self) -> None:
        if not self._current_task:
            QMessageBox.warning(self, "No Task", "Select a task first.")
            return
        if self._current_task.linked_path:
            reply = QMessageBox.question(
                self, "Existing Link",
                f"Already linked to:\n{self._current_task.linked_path}\n\nReplace it?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.No:
                return
        choice, ok = QInputDialog.getItem(
            self, "Link to…", "Link this task to:", ["A file", "A folder"], 0, False)
        if not ok:
            return
        path = (QFileDialog.getOpenFileName(self, "Select File")[0] if choice == "A file"
                else QFileDialog.getExistingDirectory(self, "Select Folder"))
        if path:
            self._current_task.linked_path = path
            self._storage.update_task(self._current_task)
            self._refresh_excel()
            QMessageBox.information(self, "Linked", f"Linked to:\n{path}")

    def _open_db(self) -> None:
        subprocess.Popen(["open", self._storage.db_path])

    def _export_and_open_excel(self) -> None:
        self._refresh_excel()
        subprocess.Popen(["open", self._excel_path])

    def _quit(self) -> None:
        self.close()

    def _refresh_excel(self) -> None:
        self._exporter.export(self._storage.get_time_entries(), self._storage.get_tasks())

    def keyPressEvent(self, event) -> None:
        if event.modifiers() == Qt.MetaModifier:
            k = event.key()
            if k == Qt.Key_B:
                self._on_start_clicked(); return
            elif k == Qt.Key_C:
                self._stop_timer(); return
            elif k == Qt.Key_M:
                self._show_menu(); return
        super().keyPressEvent(event)

    def closeEvent(self, event) -> None:
        if not self._ready:
            event.accept()
            return
        if self._mode in ("task", "task_paused"):
            self._save_current_period()
        event.accept()
        QTimer.singleShot(0, QApplication.quit)


# ── Native macOS Menu Bar & App Orchestrator ──────────────────────────────────
import objc
from AppKit import (
    NSObject, NSStatusBar, NSVariableStatusItemLength,
    NSApplication, NSApplicationActivationPolicyRegular,
)

if '_ClickHandler' not in dir():
    class _ClickHandler(NSObject):
        """Objective-C target for NSStatusBarButton clicks."""

        def init(self):
            self = objc.super(_ClickHandler, self).init()
            self._panel = None
            return self

        def setPanel_(self, panel) -> None:
            self._panel = panel

        def handleClick_(self, sender) -> None:
            if self._panel is None:
                return
            if self._panel.isMinimized():
                NSApplication.sharedApplication().activateIgnoringOtherApps_(True)
                self._panel.showNormal()
                self._panel.raise_()
                self._panel.activateWindow()
            elif self._panel.isVisible():
                self._panel.hide()
            else:
                NSApplication.sharedApplication().activateIgnoringOtherApps_(True)
                self._panel.show()
                self._panel.raise_()
                self._panel.activateWindow()


class MenuBarManager:
    """Live countdown in the native macOS menu bar."""

    def __init__(self, engine: TimerEngine, panel: MainPanel):
        self._engine = engine
        self._panel  = panel

        self._status_bar  = NSStatusBar.systemStatusBar()
        self._status_item = self._status_bar.statusItemWithLength_(
            NSVariableStatusItemLength
        )
        self._status_item.button().setTitle_("⏱  --:--")

        self._handler = _ClickHandler.alloc().init()
        self._handler.setPanel_(panel)
        btn = self._status_item.button()
        btn.setTarget_(self._handler)
        btn.setAction_("handleClick:")

        engine.tick.connect(self._on_tick)
        engine.completed.connect(self._on_completed)
        QApplication.instance().aboutToQuit.connect(self._cleanup)

    def _on_tick(self, remaining: int, _total: int) -> None:
        time_str = TimerEngine.format_time(remaining)
        if self._engine.is_break:
            self._status_item.button().setTitle_(f"☕  Break  {time_str}")
        else:
            task_name = ""
            if self._panel._current_task:
                name = self._panel._current_task.name
                task_name = (name[:16] + "…" if len(name) > 16 else name) + "  "
            self._status_item.button().setTitle_(f"⏱  {task_name}{time_str}")

    def _on_completed(self) -> None:
        if self._engine.is_break:
            self._status_item.button().setTitle_("⏱  --:--")
        else:
            self._status_item.button().setTitle_("☕  --:--")

    def _cleanup(self) -> None:
        try:
            self._status_bar.removeStatusItem_(self._status_item)
        except Exception:
            pass


class TimekeeperApp:
    """Composition root — wires all components together."""

    def __init__(self):
        self._qt = QApplication.instance() or QApplication(sys.argv)
        self._qt.setQuitOnLastWindowClosed(False)

        NSApplication.sharedApplication().setActivationPolicy_(
            NSApplicationActivationPolicyRegular
        )

        data_dir   = get_or_create_data_dir()
        db_path    = str(data_dir / "timekeeper.db")
        excel_path = str(data_dir / "timekeeper.xlsx")

        storage  = SQLiteStorage(db_path)
        exporter = ExcelExporter(excel_path)
        engine   = TimerEngine()
        panel    = MainPanel(engine, storage, exporter, excel_path)

        self._menu_bar = MenuBarManager(engine, panel)
        self._panel    = panel



    def run(self) -> None:
        self._panel.show()
        self._panel.setWindowTitle(APP_NAME)  # version shown in About, not title bar
        self._panel.raise_()
        # Get the NSWindow directly from the NSView via winId(), then
        # disable both hidesOnDeactivate and floatingPanel so the window
        # stays visible when the app is deactivated (e.g. Cmd+Tab).
        import objc
        ns_view = objc.objc_object(c_void_p=ctypes.c_void_p(int(self._panel.winId())))
        ns_window = ns_view.window()
        if ns_window is not None:
            ns_window.setHidesOnDeactivate_(False)
        self._qt.exec()


# ── Launch ────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = TimekeeperApp()
    app.run()
